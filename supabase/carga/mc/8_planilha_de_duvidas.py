# -*- coding: utf-8 -*-
"""Planilha para o Tiago responder um a um o que sobrou na raiz da Manutencao.

Uma linha por lancamento, com a nota inteira, **onde o Mais Controle lanca o
valor**, onde o resto do lancamento ja esta no ERP, e a minha sugestao. A coluna
RESPOSTA tem menu com todos os destinos validos e aceita texto livre.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = '/Users/tiagocameli/.claude/jobs/87cb0088/tmp/'
SAIDA = '/Users/tiagocameli/Downloads/raiz-manutencao-duvidas.xlsx'
MANUT = 'Manutenção/Documentação de Equipamentos'
OFIC = MANUT + ' > Oficina'

# (numero, sugestao, o que falta saber)
SUGESTAO = {
    '5080': (MANUT, 'Fica. É "transporte de equipamentos de PVH x CZS": custo da frota inteira, não de uma máquina. O MC concorda: põe estes R$ 27.000 no centro de manutenção sem etapa.'),
    '2629': ('', 'R$ 18 mil de bateria em granel. O MC também não diz máquina. Crio uma etapa "Bateria", como fiz com o Lubrificante, ou vai para a Oficina?'),
    '1683': ('Escritório Central', 'Multa da placa NXS1761, que não existe na frota. O MC põe em Empresa.'),
    '2027': ('', '"Caixa de marcha DOS CAMINHÕES", no plural. O MC também deixa sem etapa. Quais caminhões?'),
    '1536': ('', 'Não achei no MC: o único documento com esse valor fala de pá carregadeira e está a 45 dias. "03 placas das centrais dos caminhões" — quais três?'),
    '4432': ('', 'O MC põe em "Hilux Apoio - 203". Qual placa do ERP é essa? Ver a aba Hilux.'),
    '1905': ('', 'O MC põe em "Hilux de Apoio - Cinza - 209". Qual placa do ERP é essa? Ver a aba Hilux.'),
    '3500': (OFIC, 'Salário do mecânico Kennedy. O MC não tem esse documento. A Oficina é o meio-termo entre a raiz e ratear entre as máquinas.'),
    '2669': (OFIC, 'Peça sem máquina citada. O MC põe em Empresa; a Oficina mantém na manutenção.'),
    '4570': (OFIC, 'Salário. O MC não tem. O resto do lançamento já está no AC 405.'),
    '0426': (OFIC, 'Salário do Kennedy. O MC não tem.'),
    '2946': ('APLICAR O RATEIO DO MC (13 máquinas)', 'O MC rateia este frete entre 13 máquinas. Posso aplicar tudo, MENOS duas: "501 e 502 Motor Compactador de Solo" (R$ 420) não existem no cadastro do ERP. Crio, ou jogo esses R$ 420 na Oficina?'),
    '0949': ('', 'O MC põe em "Hilux Apoio - 203", a mesma do LAN-2026-4432. Ver a aba Hilux.'),
    '1468': ('Escritório Central', 'Primeira parcela do 13º. O MC põe em Empresa.'),
    '0490': (OFIC, 'Salário. O MC não tem. O resto já está no Ramal do Gama.'),
    '3732': ('', 'Radiador de qual máquina? O MC põe em Empresa, ou seja, também não sabe.'),
    '3169': ('Escritório Central', 'Frete da GOL LOG. O MC põe em Empresa.'),
    '5176': ('', '"Manutenção Equipamentos EMT" é todo o texto que existe, e o MC também deixa sem etapa. De qual máquina?'),
    '5125': ('APLICAR O RATEIO DO MC (11 destinos)', 'O MC rateia entre 11 destinos, e dois deles saem da manutenção: Casa James R$ 30 e a escola de Marechal Thaumaturgo R$ 25. Aplico assim?'),
    '2361': ('009 - Manutenção da Rodovia BR-364/AC - Lote 09 & 10', 'Frete de diesel para as obras, nada de manutenção. O MC diz 12.500 para a 009 e 2.500 para a 003; o ERP tem 11.889,67 e 2.377,93. Corrijo para o do MC?'),
    '1147': (OFIC, 'Salário. O MC não tem.'),
    '1519': ('', 'Cita a 416E E o Boiadeiro MZO 7876. Casou com um documento do MC de R$ 469,90 contra R$ 965,41 do lançamento, então rejeitei. Qual das duas 416E, e divido meio a meio com o Boiadeiro?'),
    '0632': ('', 'ESTA É A CHAVE DAS HILUX. A nota traz três placas (SQQ8F87 / QWQ1D76 / SQR1C93) e o MC divide em James R$ 220 + Tiago R$ 220 + Apoio Cinza R$ 140. Qual placa é cada uma? Ver a aba Hilux.'),
    '1891': ('', 'Cita motoniveladora 12H e retro 416E, e há duas de cada. Casou com um documento do MC que fala de serviço elétrico em caminhão, então rejeitei. Quais duas?'),
    '4307': ('APLICAR O RATEIO DO MC (9 destinos)', 'O MC rateia entre 9. Quatro deles não existem no ERP: Meloza Colorado, Caminhão Pipa Ford 101, Bobcat 18 e Carga Semi-Reboque/Prancha 104 (R$ 95 no total). Onde ponho esses?'),
    '5943': ('004 - Galpão Silo', 'O texto diz "mecânico passou 1 semana prestando serviço NO SILO". O MC não tem esse documento.'),
    '0055': (OFIC, 'Salário. O MC não tem. O resto já está na 009.'),
    '2053': ('DIVIDIR: Rolo CP56 - 01 + Rolo Pé de Carneiro CP56 - 02', '"Lavagem do rolo 01 e 02 CAT". O MC não tem esse documento. Meio a meio?'),
    '1248': ('Escritório Central', 'Frete da GOLOG. O MC põe em Empresa.'),
    '4623': ('Escritório Central', 'Frete, e o texto é literalmente "frete". O MC põe em Empresa.'),
    '1450': ('Escritório Central', 'Remendo e cola de vulcanização: borracharia, sem máquina. O MC põe em Empresa.'),
    '3288': ('DIVIDIR: Rolo CP56 - 01 + Rolo Pé de Carneiro CP56 - 02', 'Produtos para o rolo pé de carneiro 01 e 02; o rolo chapa Colorado já está na Colorado. O MC não tem esse documento.'),
    '1486': ('APLICAR O RATEIO DO MC', 'O MC diz Colorado R$ 319,48 + Caminhão Pipa L1318/50 MZO-4486 - 02 R$ 65,52. O ERP tem R$ 296,32 na raiz e só R$ 88,68 na Colorado, ou seja, está errado nos dois lados.'),
    '5713': ('Escritório Central', 'Multa da placa QLZ6A95. O MC põe em Empresa. Essa placa está na frota?'),
    '3614': ('Escritório Central', '"Importado do maiscontrole" é todo o texto, e no MC a descrição é vazia. O MC põe em Empresa.'),
    '2467': ('', 'Pino da balança e deslizante GUERRA: é uma das carretas Guerra. O MC também deixa sem etapa. B2D095 - 02 ou B2T093 - 03?'),
    '4714': (MANUT, 'Frete de peça da VEMAP. O MC não tem esse documento.'),
    '0793': ('APLICAR O RATEIO DO MC (7 destinos)', 'O MC rateia entre 7. Dois não encaixam: "Skidy" R$ 40 não existe no ERP, e R$ 40 vão para Despesas Diversas da 009.'),
    '2048': ('DIVIDIR: Rolo Chapa CB10 - 01 + Rolo de Pneu CW34 - 01', '"Ar condicionado do rolo chapa e rolo de pneu". Casou com um documento do MC que fala de passagem CZS x Marechal, então rejeitei. Meio a meio?'),
    '3369': (OFIC, 'Material elétrico e de solda: consumo de oficina. O MC não tem esse documento.'),
    '2594': ('Escritório Central', 'Licenciamento da placa NXS2265, que não está na frota. O MC põe em Empresa.'),
    '4535': ('Escritório Central', 'Licenciamento da placa QLW3B35, que não está na frota. O MC põe em Empresa.'),
    '0115': ('Escritório Central', 'Licenciamento da placa NXS3504, que não está na frota. O MC põe em Empresa.'),
    '1635': ('Escritório Central', 'Licenciamento da placa QLZ3923, que não está na frota. O MC põe em Empresa.'),
    '0702': ('', 'Disco, fita e broca "para rolo pé de carneiro CP56", e há dois CP56. O MC não tem esse documento. Qual, ou meio a meio?'),
    '0373': ('003 - Recuperação do Ramal do Gama', 'Despesa de saída para o Ramal do Gama e AC 405. O MC diz 450 para a 003 e 550 para a 007; o ERP tem 525,21 e 270,26.'),
    '5018': ('Escritório Central', 'Licenciamento da QLZ3923, mesma placa do LAN-2026-1635.'),
    '1769': ('Escritório Central', 'Licenciamento da NXS2265, mesma placa do LAN-2026-2594.'),
    '3176': (OFIC, '"DOLA - MANGUEIRA". Dola é o mecânico; mangueira é consumo de oficina.'),
    '2686': ('', 'O MC põe em "Hilux de Apoio - Cinza - 209", a mesma do LAN-2026-1905. Bate com o texto, que diz "Hilux CINZA". Ver a aba Hilux.'),
    '0295': (MANUT + ' > Trator de Esteira D6NXL - 01', 'Arruelas, parafusos e porcas "D6N": é o D6NXL.'),
    '3137': (OFIC, 'Salário. Casou com um documento do MC de R$ 2.297,77 contra R$ 1.582,70, então rejeitei. O resto já está no AC 405.'),
    '4470': ('Escritório Central', 'IPVA da NXS2265, que não está na frota.'),
    '5951': ('', 'Abastecimento de R$ 50 num "caixa do dia". Casou com gasolina da BROS 160 no MC, mas por CNPJ e a 4 dias, então rejeitei. De qual veículo?'),
    '2377': ('', 'Bucha estabilizadora MB1720/1938: é modelo de caminhão. O MC não tem. Qual da frota?'),
    '5952': ('', '"Caixa do dia" de R$ 40 na Cruzeiro Peças. O MC não tem. De qual máquina?'),
    '5847': ('', 'Frete da solenoide "da pá carregadeira". O MC não tem. Há a 924K - 01, a W20 - 02 e a Komatsu 150. Qual?'),
    '5937': ('', '"Caixa do dia" de R$ 14 na Cruzeiro Peças. O MC não tem. De qual máquina?'),
    '5938': (MANUT, 'Frete de R$ 10 no "caixa do dia". O MC não tem.'),
}

linhas = [l for l in io.open(BASE + 'duvidas.txt', encoding='utf-8').read().split('\n') if l.strip()]
destinos = [d for d in io.open(BASE + 'destinos.txt', encoding='utf-8').read().split('\n') if d.strip()]
NO_MC = {}
for l in io.open(BASE + 'mc_dos_59.txt', encoding='utf-8').read().split('\n'):
    if l.strip():
        n, t = l.split('§', 1)
        NO_MC[n] = t

wb = Workbook()
CAB = PatternFill('solid', fgColor='45464B')
FUNDO_RESP = PatternFill('solid', fgColor='FBF1E3')
FUNDO_SUG = PatternFill('solid', fgColor='F7F7F5')
FUNDO_MC = PatternFill('solid', fgColor='EDF3EE')
borda = Border(*[Side(style='thin', color='D8D5CE')] * 4)

# ---------------------------------------------------------------- aba Duvidas
ws = wb.active
ws.title = 'Dúvidas'
COLS = [('#', 5), ('Lançamento', 15), ('Data', 11), ('Valor na raiz', 14),
        ('Total do lanç.', 14), ('Fornecedor', 26), ('Descrição da nota', 56),
        ('Categoria', 18), ('ONDE ESTÁ NO MAIS CONTROLE', 68),
        ('Onde o resto do lanç. já está no ERP', 40),
        ('Minha sugestão', 40), ('O que falta saber', 62), ('SUA RESPOSTA', 38)]
NC = len(COLS)
COL_MC, COL_SUG, COL_FALTA, COL_RESP = 9, 11, 12, 13

ws.append([c for c, _ in COLS])
for i, (_, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cel = ws.cell(row=1, column=i)
    cel.font = Font(bold=True, color='FFFFFF', size=10)
    cel.fill = CAB
    cel.alignment = Alignment(vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 32

total = 0.0
for n, linha in enumerate(linhas, 1):
    p = linha.split('§')
    num, dt, fatia, valor, forn, desc, cat = p[0], p[1], p[2], p[3], p[4], p[5], p[6]
    resto = p[8] if len(p) > 8 else ''
    curto = num.replace('LAN-2026-', '')
    sug, falta = SUGESTAO.get(curto, ('', ''))
    total += float(fatia)
    ws.append([n, num, dt, float(fatia), float(valor), forn, desc, cat,
               NO_MC.get(curto, '(não procurei)'),
               resto or '(o lançamento inteiro está na raiz)', sug, falta, ''])
    r = ws.max_row
    for c in range(1, NC + 1):
        cel = ws.cell(row=r, column=c)
        cel.border = borda
        cel.alignment = Alignment(vertical='top',
                                  wrap_text=(c in (7, COL_MC, 10, COL_SUG, COL_FALTA)))
        if c in (4, 5):
            cel.number_format = 'R$ #,##0.00'
        if c == COL_MC:
            cel.fill = FUNDO_MC
            cel.font = Font(size=9)
        if c in (COL_SUG, COL_FALTA):
            cel.fill = FUNDO_SUG
            cel.font = Font(size=9, color='6B6B6B')
        if c == COL_RESP:
            cel.fill = FUNDO_RESP
        if c in (1, 3):
            cel.alignment = Alignment(vertical='top', horizontal='center')
        if c == 7:
            cel.font = Font(size=9)
    ws.row_dimensions[r].height = 58

r = ws.max_row + 1
ws.cell(row=r, column=3, value='TOTAL').font = Font(bold=True)
c = ws.cell(row=r, column=4, value=round(total, 2))
c.font = Font(bold=True)
c.number_format = 'R$ #,##0.00'
ws.cell(row=r, column=6, value='%d lançamentos' % len(linhas)).font = Font(bold=True)
ws.freeze_panes = 'C2'
ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(NC), ws.max_row - 1)

# ---------------------------------------------------------------- aba Hilux
wh = wb.create_sheet('Hilux')
wh.append(['As quatro Hilux do MC contra as cinco do ERP — R$ 10.864,38'])
wh.cell(row=1, column=1).font = Font(bold=True, size=12)
wh.append([])
for t in ['O MC lança em quatro nomes. O ERP tem cinco Hilux, todas com placa. Dizendo',
          'qual placa é cada nome do MC, cinco linhas da planilha se resolvem de uma vez.',
          '']:
    wh.append([t])
wh.append(['Nome no Mais Controle', 'Valor', 'Aparece em', 'Qual placa do ERP é essa?'])
r0 = wh.max_row
for i in range(1, 5):
    cel = wh.cell(row=r0, column=i)
    cel.font = Font(bold=True, color='FFFFFF', size=10)
    cel.fill = CAB
HIL = [
    ('Hilux Apoio - 203 - Peças', 6970.00, 'LAN-2026-4432 (R$ 5.450,00) e LAN-2026-0949 (R$ 1.520,00)'),
    ('Hilux de Apoio - Cinza - 209 - Peças', 3454.38, 'LAN-2026-1905 (R$ 3.194,38), LAN-2026-2686 (R$ 120,00) e R$ 140,00 do LAN-2026-0632'),
    ('0205 Hilux James', 220.00, 'R$ 220,00 do LAN-2026-0632'),
    ('0207 Hilux Tiago', 220.00, 'R$ 220,00 do LAN-2026-0632'),
]
for nome, val, onde in HIL:
    wh.append([nome, val, onde, ''])
    r = wh.max_row
    wh.cell(row=r, column=2).number_format = 'R$ #,##0.00'
    wh.cell(row=r, column=4).fill = FUNDO_RESP
    for i in range(1, 5):
        wh.cell(row=r, column=i).border = borda
        wh.cell(row=r, column=i).alignment = Alignment(vertical='top', wrap_text=(i == 3))
    wh.row_dimensions[r].height = 30
r = wh.max_row + 1
wh.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
c = wh.cell(row=r, column=2, value=10864.38)
c.font = Font(bold=True)
c.number_format = 'R$ #,##0.00'
wh.append([])
for t in ['As cinco Hilux do ERP:',
          '   Hilux CDLOWA4SD SQQ-8F87 - 06',
          '   Hilux CDSRVA4FD QWQ-1D76 - 05',
          '   Hilux CDSRXA4FD QLY-7H84 - 04',
          '   Hilux CHLSTM4FD QWQ-3H97 - 01',
          '   Hilux SQR1C93 - 07',
          '',
          'Pista: a nota do LAN-2026-0632 traz TRÊS placas — SQQ8F87, QWQ1D76 e SQR1C93 —',
          'e o MC divide esse mesmo documento entre James, Tiago e Apoio Cinza. Então essas',
          'três placas são esses três nomes, em alguma ordem. Sobram QLY-7H84 e QWQ-3H97,',
          'e uma delas deve ser a "Apoio 203", que é de longe a mais cara das quatro.']:
    wh.append([t])
wh.column_dimensions['A'].width = 40
wh.column_dimensions['B'].width = 14
wh.column_dimensions['C'].width = 62
wh.column_dimensions['D'].width = 40

# ---------------------------------------------------------------- aba Destinos
wd = wb.create_sheet('Destinos')
wd.append(['Destinos válidos (o menu da coluna SUA RESPOSTA lê esta lista)'])
wd.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF')
wd.cell(row=1, column=1).fill = CAB
wd.column_dimensions['A'].width = 70
for d in destinos:
    wd.append([d])
wd.freeze_panes = 'A2'

dv = DataValidation(type='list', formula1='=Destinos!$A$2:$A$%d' % (len(destinos) + 1),
                    allow_blank=True, showDropDown=False)
dv.errorStyle = 'warning'
dv.promptTitle = 'Destino'
dv.prompt = ('Escolha da lista ou escreva livre. Para dividir: '
             '"DIVIDIR: <destino> + <destino>". Para aceitar o rateio do MC: '
             '"MC". Para deixar como está: "FICA".')
dv.showInputMessage = True
ws.add_data_validation(dv)
dv.add('%s2:%s%d' % (get_column_letter(COL_RESP), get_column_letter(COL_RESP), len(linhas) + 1))

# ---------------------------------------------------------------- aba Como usar
wi = wb.create_sheet('Como usar', 0)
wi.column_dimensions['A'].width = 104
brl = '{:,.2f}'.format(total).replace(',', '@').replace('.', ',').replace('@', '.')
texto = [
    ('Raiz da Manutenção: o que ainda está em dúvida', True),
    ('', False),
    ('São %d lançamentos e R$ %s ainda no centro "Manutenção/Documentação de' % (len(linhas), brl), False),
    ('Equipamentos" sem etapa — ou seja, sem máquina.', False),
    ('', False),
    ('Preencha só a coluna "SUA RESPOSTA" (a última, de fundo âmbar).', True),
    ('', False),
    ('   • A célula tem menu com os 81 destinos do cadastro, e aceita texto livre.', False),
    ('   • Para dividir:      DIVIDIR: Rolo CP56 - 01 + Rolo Pé de Carneiro CP56 - 02', False),
    ('     (divido meio a meio; se a proporção for outra, escreva os valores)', False),
    ('   • Para aceitar o rateio do Mais Controle daquela linha, escreva só:  MC', False),
    ('   • Para deixar onde está:  FICA', False),
    ('', False),
    ('A coluna verde "ONDE ESTÁ NO MAIS CONTROLE" é o dado mais forte da planilha:', True),
    ('é o rateio real do documento no MC, com centro, etapa e valor de cada fatia.', False),
    ('Ela diz também quando eu casei o documento e depois rejeitei o casamento, e por quê,', False),
    ('e quando o MC simplesmente não tem aquele documento.', False),
    ('', False),
    ('   38 dos 59 casaram com um documento do MC', False),
    ('    6 casaram e eu rejeitei (o valor ou o assunto não fecha) — está escrito qual e por quê', False),
    ('   15 o MC não tem, e desses 6 são salário de mecânico', False),
    ('', False),
    ('Em 5 linhas a sugestão é "APLICAR O RATEIO DO MC": nessas o MC divide entre várias', False),
    ('máquinas e eu posso aplicar tudo, mas a coluna ao lado avisa quais fatias caem em', False),
    ('máquina que não existe no cadastro do ERP.', False),
    ('', False),
    ('A aba Hilux resolve R$ 10.864,38 de uma vez: são cinco linhas da planilha que', True),
    ('dependem de saber qual placa do ERP é cada um dos quatro nomes do MC.', False),
    ('', False),
    ('Não está aqui: a compra de lubrificante em granel (R$ 77.184,64), que já foi para a', False),
    ('etapa Lubrificante nova. Nem as 38 trocas de óleo que citam a máquina — essas ficam', False),
    ('na máquina, porque mover destruiria o custo por equipamento.', False),
]
for t, negrito in texto:
    wi.append([t])
    if negrito:
        wi.cell(row=wi.max_row, column=1).font = Font(bold=True, size=12 if wi.max_row == 1 else 11)

wb.save(SAIDA)
print('gravei %s' % SAIDA)
print('%d linhas, R$ %.2f | %d destinos' % (len(linhas), total, len(destinos)))
print('com coluna do MC preenchida: %d' % sum(1 for l in linhas
                                              if l.split('§')[0].replace('LAN-2026-', '') in NO_MC))
